import openpyxl

wb = openpyxl.load_workbook(r'd:\TimeT\ALL.xlsx', data_only=True)

for sheet_name in ['AA', 'BA']:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*60}")
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx > 50:
            print(f"... (truncated, max row: {ws.max_row})")
            break
        vals = [str(c) if c is not None else '' for c in row]
        if any(vals):
            print(f"Row {row_idx}: {' | '.join(vals)}")
